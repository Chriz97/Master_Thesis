import os
import pandas as pd
import datetime
import openpyxl
import copy
from openpyxl.styles import Alignment

benchmark_file = r"C:\Users\Christoph\Documents\Master Thesis\Performance\Benchmark.txt"

# Define the function to read the benchmark file and create a DataFrame
def read_file_and_create_dataframe(file_path):
    benchmark_data = []
    game_names = set()  # Use a set to avoid duplicate game names
    with open(file_path, 'r') as file:
        lines = file.readlines()
        i = 0
        while i < len(lines):
            if "benchmark completed" in lines[i]:
                benchmark_info = extract_benchmark_info(lines[i:i+7])
                benchmark_data.append(benchmark_info)
                game_names.add(benchmark_info["Game Name"])  # Add game name to the set
            i += 1
    df = pd.DataFrame(benchmark_data)
    return df, game_names


# Define the function to extract benchmark information
def extract_benchmark_info(file_lines):
    # Parsing the first line for game name, date, and time
    game_name_line = file_lines[0]
    parts = game_name_line.split(',')
    duration = file_lines[0].split()[-2]+"s"
    date_part= parts[0].strip()
    time_part = file_lines[0].split()[1]
    game_name_part = parts[1].split()[1].strip(".exe")

    # Correctly parsing the benchmark metrics
    average_frame_rate = float(file_lines[1].split(':')[1].strip().split(' ')[0])
    minimum_frame_rate = float(file_lines[2].split(':')[1].strip().split(' ')[0])
    maximum_frame_rate = float(file_lines[3].split(':')[1].strip().split(' ')[0])
    low_1_percent = float(file_lines[4].split(':')[1].strip().split(' ')[0])
    low_0_1_percent = float(file_lines[5].split(':')[1].strip().split(' ')[0])

    return {
        "Game Name": game_name_part,
        "Duration" : duration,
        "Date": date_part,
        "Time": time_part,
        "Average FPS": average_frame_rate,
        "Minimum FPS": minimum_frame_rate,
        "Maximum FPS": maximum_frame_rate,
        "1% low": low_1_percent,
        "0.1% low": low_0_1_percent
    }


# Read the benchmark file and create DataFrame
df, game_names = read_file_and_create_dataframe(benchmark_file)
game_names = list(game_names)[0]  # Extract the game name from the set
now = datetime.datetime.now()
timestamp = now.strftime("%m_%d_%H_%M")  # Example: "10_30_11_44"


# Function to get user inputs for benchmark information
def check_benchmark():
    technologies = ["Native", "DLSS", "XeSS", "FSR"]
    upscaling_setting = ["None", "Performance", "Balanced", "Quality"]
    graphics_settings = ["Low", "Medium", "High"]
    resolution = ["1080p", "1440p"]
    graphics_card = ["3060", "4060"]
    tech_index = int(input(f"Select the technology ({', '.join([f'{i}: {tech}' for i, tech
                                                                in enumerate(technologies)])}): "))
    upscaling_index = int(input(f"Select the upscaling setting ({', '.join([f'{i}: {setting}' for i, setting
                                                                            in enumerate(upscaling_setting)])}): "))
    graphics_index = int(input(f"Select the graphics setting ({', '.join([f'{i}: {setting}' for i, setting
                                                                          in enumerate(graphics_settings)])}): "))
    resolution_index = int(input(f"Select the resolution ({', '.join([f'{i}: {res}' for i, res
                                                                      in enumerate(resolution)])}): "))
    graphics_card_index = int(input(f"Select the graphics card ({', '.join([f'{i}: {gpu}' for i, gpu
                                                                      in enumerate(graphics_card)])}): "))
    technology_name = technologies[tech_index]
    upscaling_name = upscaling_setting[upscaling_index]
    graphics_name = graphics_settings[graphics_index]
    resolution_name = resolution[resolution_index]
    graphics_card_name = graphics_card[graphics_card_index]
    return technology_name, upscaling_name, graphics_name, resolution_name, graphics_card_name


# Get user inputs for benchmark information
technology_name, upscaling_name, graphics_name, resolution_name, graphics_card_name = check_benchmark()

# Define the target directory and file name
output_directory = r"C:\Users\Christoph\Documents\Master Thesis\Benchmark"

# Now include game_name in your output file name
output_file_name = f"benchmark_{game_names}_{timestamp}_{technology_name}_{upscaling_name}_{graphics_name}_{resolution_name}_{graphics_card_name}.xlsx"

# Full file path
output_file_path = os.path.join(output_directory, output_file_name)

# Save the DataFrame to Excel
df.to_excel(output_file_path, index=False)

print(f"Excel file '{output_file_name}' has been created.")

# Open the created Excel file and add cells from "Book1.xlsx"
wb = openpyxl.load_workbook(filename=output_file_path)
sheet = wb.active

# Open "Book1.xlsx" and get the range of cells to copy
source_wb = openpyxl.load_workbook(filename=r"C:\Users\Christoph\Documents\Master Thesis\Book1.xlsx")
source_sheet = source_wb.active
cell_range = source_sheet["A5:J21"]

# Copy cells with values and formats
start_row = sheet.max_row + 2
for row in cell_range:
    col_num = 1
    for cell in row:
        target_cell = sheet.cell(row=start_row, column=col_num)
        target_cell.value = cell.value

        # Copy cell formatting
        target_cell.number_format = cell.number_format

        # Copy font properties
        target_cell.font = copy.copy(cell.font)

        # Copy alignment properties
        target_cell.alignment = copy.copy(cell.alignment)

        # Copy border properties
        target_cell.border = copy.copy(cell.border)

        # Copy fill properties
        target_cell.fill = copy.copy(cell.fill)

        col_num += 1
    start_row += 1

# Merge and center specified cells
cells_to_merge = ["A5:D5", "A7:I7", "A14:F14", "A16:F16", "A21:F21"]
for cell_range in cells_to_merge:
    sheet.merge_cells(cell_range)
    # Center text in merged cells
    for row in sheet[cell_range]:
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

# Adjust the width of columns A and J
sheet.column_dimensions['A'].width = 16.57
sheet.column_dimensions['J'].width = 11.14

# Save the updated Excel file
wb.save(output_file_path)

print(f"Cells from 'Book1.xlsx' added to '{output_file_name}'.")

# Last step: remove the original file
if os.path.exists(benchmark_file):
    os.remove(benchmark_file)
    print("Benchmark.txt has been deleted.")
else:
    print("The file does not exist.")
